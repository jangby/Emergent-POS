from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import logging
import uuid
import base64
import hashlib
import hmac
import asyncio
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

import bcrypt
import jwt
import httpx
from cryptography.fernet import Fernet
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Form
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import JSONResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

# --- Config ---
MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
JWT_SECRET = os.environ.get('JWT_SECRET', 'change-me')
JWT_ALGO = 'HS256'
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'admin@example.com')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')
SETTINGS_ENCRYPTION_KEY = os.environ.get('SETTINGS_ENCRYPTION_KEY', '').encode()

fernet = Fernet(SETTINGS_ENCRYPTION_KEY) if SETTINGS_ENCRYPTION_KEY else None

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="KasirPintar AI")
api = APIRouter(prefix="/api")

# --- Utils ---
def now_utc():
    return datetime.now(timezone.utc)

def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False

def create_access_token(uid: str, email: str) -> str:
    return jwt.encode({"sub": uid, "email": email, "type": "access",
                       "exp": now_utc() + timedelta(hours=8)}, JWT_SECRET, algorithm=JWT_ALGO)

def create_refresh_token(uid: str) -> str:
    return jwt.encode({"sub": uid, "type": "refresh",
                       "exp": now_utc() + timedelta(days=30)}, JWT_SECRET, algorithm=JWT_ALGO)

def set_auth_cookies(resp: Response, access: str, refresh: str):
    resp.set_cookie("access_token", access, httponly=True, secure=True, samesite="none",
                    max_age=28800, path="/")
    resp.set_cookie("refresh_token", refresh, httponly=True, secure=True, samesite="none",
                    max_age=2592000, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        if payload.get("type") != "access":
            raise HTTPException(401, "Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(401, "User not found")
        # Multi-tenant: tenant_id is either the stored one, or the user's own id (self-tenant for owners)
        user["tenant_id"] = user.get("tenant_id") or user["id"]
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

def tenant_of(user: dict) -> str:
    """Return the tenant id for the given authenticated user."""
    return user.get("tenant_id") or user["id"]


# --- Gemini (BYOK per-tenant) helpers ---
GEMINI_MODEL_TEXT = "gemini-3.6-flash"
GEMINI_MODEL_VISION = "gemini-3.6-flash"
GEMINI_TIMEOUT_SEC = 45  # keep under Cloudflare / ingress 60-100s cutoff

# genai.configure() sets a *process-global* API key. Serialize configure+call so
# concurrent tenants with different keys can't clobber each other mid-call.
_gemini_lock = asyncio.Lock()

async def get_gemini_key(tenant_id: str) -> str:
    doc = await db.settings.find_one({"tenant_id": tenant_id}) or {}
    enc_key = doc.get("gemini", {}).get("api_key_enc")
    if not enc_key:
        raise HTTPException(
            400,
            "Gemini API Key belum dikonfigurasi. Silakan atur di menu Pengaturan → Konfigurasi AI."
        )
    return dec(enc_key)

def _extract_text_safely(resp) -> str:
    """`resp.text` raises when candidates are blocked or empty. Recover gracefully."""
    try:
        return resp.text or ""
    except Exception:
        pass
    try:
        pieces = []
        for cand in getattr(resp, "candidates", []) or []:
            content = getattr(cand, "content", None)
            for part in (getattr(content, "parts", []) or []):
                t = getattr(part, "text", None)
                if t:
                    pieces.append(t)
        return "".join(pieces)
    except Exception:
        return ""

async def gemini_generate(tenant_id: str, system: str, user_prompt: str,
                          image_b64: Optional[str] = None,
                          image_mime: str = "image/jpeg",
                          model_name: Optional[str] = None) -> str:
    """Call Gemini via the official google-generativeai SDK using the tenant's own API key.

    Guarantees:
      - Never blocks the FastAPI event loop (runs in threadpool).
      - Never crashes: any SDK exception is translated to an HTTPException with a
        clear message and the full traceback is written to server logs.
      - Enforces a hard timeout ({GEMINI_TIMEOUT_SEC}s) so Cloudflare / ingress
        never sees an idle request.
    """
    key = await get_gemini_key(tenant_id)
    mname = model_name or (GEMINI_MODEL_VISION if image_b64 else GEMINI_MODEL_TEXT)

    def _run():
        import google.generativeai as genai
        genai.configure(api_key=key)
        model = genai.GenerativeModel(model_name=mname, system_instruction=system)
        parts: list = [user_prompt]
        if image_b64:
            parts.append({"mime_type": image_mime, "data": base64.b64decode(image_b64)})
        resp = model.generate_content(
            parts,
            generation_config={"response_mime_type": "text/plain"},
            request_options={"timeout": GEMINI_TIMEOUT_SEC},
        )
        return _extract_text_safely(resp)

    try:
        async with _gemini_lock:
            # asyncio.wait_for adds an outer safety net in case the SDK ignores its timeout
            return await asyncio.wait_for(
                run_in_threadpool(_run),
                timeout=GEMINI_TIMEOUT_SEC + 10,
            )
    except HTTPException:
        raise
    except asyncio.TimeoutError:
        logging.error("Gemini call timed out after %ss", GEMINI_TIMEOUT_SEC + 10)
        raise HTTPException(
            504,
            "Gemini terlalu lama merespons. Coba lagi sebentar atau periksa koneksi internet Anda."
        )
    except Exception as e:
        # Always log full traceback for debugging.
        logging.error("Gemini call failed: %s", e, exc_info=True)
        msg = str(e) or type(e).__name__
        low = msg.lower()
        if ("api key" in low or "api_key" in low or "api_key_invalid" in low
                or "unauthenticated" in low or "401" in low):
            raise HTTPException(
                400,
                f"Gemini API Key tidak valid atau ditolak. Perbarui key di Pengaturan → Konfigurasi AI. ({msg[:200]})"
            )
        if "permission" in low or "403" in low:
            raise HTTPException(
                403,
                f"Gemini menolak permintaan (permission denied). Pastikan API key Anda mengaktifkan Generative Language API. ({msg[:200]})"
            )
        if "quota" in low or "resource_exhausted" in low or "429" in low or "rate" in low:
            raise HTTPException(
                429,
                f"Kuota Gemini terlampaui. Coba lagi nanti atau upgrade paket API Anda. ({msg[:200]})"
            )
        if "deadline" in low or "timeout" in low or "504" in low:
            raise HTTPException(
                504,
                f"Gemini timeout. Coba lagi sebentar. ({msg[:200]})"
            )
        if "safety" in low or "blocked" in low or "block_reason" in low:
            raise HTTPException(
                400,
                f"Permintaan diblokir oleh filter keamanan Gemini. Coba ubah kata pada prompt. ({msg[:200]})"
            )
        raise HTTPException(
            500,
            f"Gemini gagal merespons. Silakan coba lagi. ({msg[:200]})"
        )

def enc(v: str) -> str:
    return fernet.encrypt(v.encode()).decode() if fernet and v else v

def dec(v: str) -> str:
    if not fernet or not v:
        return v
    try:
        return fernet.decrypt(v.encode()).decode()
    except Exception:
        return ""

def strip_id(d: dict) -> dict:
    if d and "_id" in d:
        d.pop("_id", None)
    return d

# --- Auth models ---
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = "Merchant"

class LoginIn(BaseModel):
    email: EmailStr
    password: str

# --- Auth routes ---
@api.post("/auth/register")
async def register(body: RegisterIn, response: Response):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already registered")
    uid = str(uuid.uuid4())
    # Multi-tenant: new owner IS their own tenant
    user = {"id": uid, "email": email, "name": body.name, "role": "owner",
            "tenant_id": uid,
            "password_hash": hash_password(body.password),
            "created_at": now_utc().isoformat()}
    await db.users.insert_one(user)
    # Seed a default settings doc for the new tenant
    await db.settings.update_one(
        {"tenant_id": uid},
        {"$setOnInsert": {"tenant_id": uid,
                          "store": StoreProfile(name=f"Toko {body.name}").model_dump(),
                          "midtrans": {}}},
        upsert=True,
    )
    access = create_access_token(uid, email)
    refresh = create_refresh_token(uid)
    set_auth_cookies(response, access, refresh)
    return {"id": uid, "email": email, "name": body.name, "role": "owner",
            "access_token": access}

@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "Email atau password salah")
    access = create_access_token(user["id"], email)
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    return {"id": user["id"], "email": email, "name": user["name"], "role": user.get("role", "owner"),
            "access_token": access}

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

@api.post("/auth/refresh")
async def refresh_token_ep(request: Request, response: Response):
    rt = request.cookies.get("refresh_token")
    if not rt:
        raise HTTPException(401, "No refresh token")
    try:
        payload = jwt.decode(rt, JWT_SECRET, algorithms=[JWT_ALGO])
        if payload.get("type") != "refresh":
            raise HTTPException(401, "Invalid token")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(401, "User not found")
        access = create_access_token(user["id"], user["email"])
        response.set_cookie("access_token", access, httponly=True, secure=True,
                            samesite="none", max_age=28800, path="/")
        return {"access_token": access}
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

# --- Product models ---
class ProductIn(BaseModel):
    name: str
    category: str
    stock: int = 0
    buy_price: int = 0
    sell_price: int = 0
    sku: Optional[str] = None
    image_url: Optional[str] = None

# --- Products ---
@api.get("/products")
async def list_products(user: dict = Depends(get_current_user)):
    items = await db.products.find({"tenant_id": tenant_of(user)}, {"_id": 0}).sort("name", 1).to_list(1000)
    return items

@api.post("/products")
async def create_product(body: ProductIn, user: dict = Depends(get_current_user)):
    pid = str(uuid.uuid4())
    doc = {"id": pid, "tenant_id": tenant_of(user), **body.model_dump(),
           "created_at": now_utc().isoformat()}
    await db.products.insert_one(doc)
    return strip_id(doc)

@api.put("/products/{pid}")
async def update_product(pid: str, body: ProductIn, user: dict = Depends(get_current_user)):
    r = await db.products.update_one({"id": pid, "tenant_id": tenant_of(user)},
                                     {"$set": body.model_dump()})
    if r.matched_count == 0:
        raise HTTPException(404, "Not found")
    doc = await db.products.find_one({"id": pid, "tenant_id": tenant_of(user)}, {"_id": 0})
    return doc

@api.delete("/products/{pid}")
async def delete_product(pid: str, user: dict = Depends(get_current_user)):
    await db.products.delete_one({"id": pid, "tenant_id": tenant_of(user)})
    return {"ok": True}

@api.get("/products/categories")
async def categories(user: dict = Depends(get_current_user)):
    cats = await db.products.distinct("category", {"tenant_id": tenant_of(user)})
    return sorted(cats)

# --- Transactions ---
class CartItem(BaseModel):
    product_id: str
    name: str
    qty: int
    price: int  # sell price at time
    buy_price: int = 0
    discount: int = 0  # per line in Rp

class TransactionIn(BaseModel):
    items: List[CartItem]
    subtotal: int
    discount: int = 0
    tax: int = 0
    total: int
    payment_method: Literal["cash", "qris", "credit"]
    cash_tendered: int = 0
    change: int = 0
    qris_order_id: Optional[str] = None
    qris_status: Optional[str] = None
    applied_promos: List[dict] = []
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None

@api.post("/transactions")
async def create_transaction(body: TransactionIn, user: dict = Depends(get_current_user)):
    tid = str(uuid.uuid4())
    order_id = f"TRX-{datetime.now().strftime('%y%m%d%H%M%S')}-{tid[:4].upper()}"
    net_profit = 0
    for it in body.items:
        net_profit += (it.price - it.buy_price) * it.qty - it.discount
    # Link to open shift, if any
    open_shift = await db.shifts.find_one({"user_id": user["id"], "status": "open",
                                           "tenant_id": tenant_of(user)})
    is_credit = body.payment_method == "credit"
    status = "completed"
    if is_credit:
        status = "unpaid"
    elif body.payment_method == "qris" and body.qris_status not in ("settlement", "capture"):
        status = "pending"
    doc = {"id": tid, "tenant_id": tenant_of(user),
           "order_id": order_id,
           "items": [i.model_dump() for i in body.items],
           "subtotal": body.subtotal, "discount": body.discount,
           "tax": body.tax, "total": body.total,
           "payment_method": body.payment_method,
           "cash_tendered": body.cash_tendered, "change": body.change,
           "qris_order_id": body.qris_order_id, "qris_status": body.qris_status,
           "applied_promos": body.applied_promos,
           "status": status,
           "net_profit": net_profit,
           "cashier": user["email"],
           "cashier_id": user["id"],
           "shift_id": open_shift["id"] if open_shift else None,
           "customer_id": body.customer_id,
           "customer_name": body.customer_name,
           "customer_phone": body.customer_phone,
           "created_at": now_utc().isoformat()}
    await db.transactions.insert_one(doc)

    # Deduct stock (tenant-scoped)
    pids = []
    for it in body.items:
        await db.products.update_one({"id": it.product_id, "tenant_id": tenant_of(user)},
                                     {"$inc": {"stock": -it.qty}})
        pids.append(it.product_id)

    # Credit: create/upsert customer and add debt ledger
    if is_credit and body.customer_phone:
        cust_id = body.customer_id
        if not cust_id:
            existing = await db.customers.find_one({"phone": body.customer_phone,
                                                    "tenant_id": tenant_of(user)})
            if existing:
                cust_id = existing["id"]
            else:
                cust_id = str(uuid.uuid4())
                await db.customers.insert_one({
                    "id": cust_id, "tenant_id": tenant_of(user),
                    "name": body.customer_name or body.customer_phone,
                    "phone": body.customer_phone, "total_debt": 0,
                    "created_at": now_utc().isoformat()
                })
        await db.customers.update_one({"id": cust_id, "tenant_id": tenant_of(user)},
                                      {"$inc": {"total_debt": body.total}})
        await db.debt_ledger.insert_one({
            "id": str(uuid.uuid4()), "tenant_id": tenant_of(user),
            "customer_id": cust_id,
            "transaction_id": tid, "order_id": order_id,
            "type": "debit", "amount": body.total,
            "created_at": now_utc().isoformat(),
        })
        doc["customer_id"] = cust_id

    # Low-stock WA alert (fire and forget)
    try:
        asyncio.create_task(check_low_stock_and_alert(tenant_of(user), pids))
    except Exception:
        pass

    return strip_id(doc)

@api.get("/transactions")
async def list_transactions(q: Optional[str] = None, days: int = 30,
                            user: dict = Depends(get_current_user)):
    cutoff = (now_utc() - timedelta(days=days)).isoformat()
    query = {"tenant_id": tenant_of(user), "created_at": {"$gte": cutoff}}
    if q:
        query["order_id"] = {"$regex": q, "$options": "i"}
    items = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items

@api.get("/transactions/{tid}")
async def get_transaction(tid: str, user: dict = Depends(get_current_user)):
    doc = await db.transactions.find_one({"id": tid, "tenant_id": tenant_of(user)}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    return doc

# --- Analytics ---
@api.get("/analytics/summary")
async def analytics_summary(user: dict = Depends(get_current_user)):
    all_tx = await db.transactions.find({"tenant_id": tenant_of(user),
                                         "status": "completed"}, {"_id": 0}).to_list(5000)
    today = now_utc().date()
    week_ago = today - timedelta(days=6)
    month_ago = today - timedelta(days=29)

    def tx_date(t):
        return datetime.fromisoformat(t["created_at"]).date()

    total_today = sum(t["total"] for t in all_tx if tx_date(t) == today)
    total_week = sum(t["total"] for t in all_tx if tx_date(t) >= week_ago)
    total_month = sum(t["total"] for t in all_tx if tx_date(t) >= month_ago)
    profit_month = sum(t.get("net_profit", 0) for t in all_tx if tx_date(t) >= month_ago)

    # Daily revenue vs profit (last 30 days)
    daily = {}
    for i in range(30):
        d = today - timedelta(days=29 - i)
        daily[d.isoformat()] = {"date": d.isoformat(), "revenue": 0, "profit": 0}
    for t in all_tx:
        d = tx_date(t).isoformat()
        if d in daily:
            daily[d]["revenue"] += t["total"]
            daily[d]["profit"] += t.get("net_profit", 0)
    daily_series = list(daily.values())

    # Top products
    top = {}
    for t in all_tx:
        for it in t["items"]:
            top.setdefault(it["name"], {"name": it["name"], "qty": 0, "revenue": 0})
            top[it["name"]]["qty"] += it["qty"]
            top[it["name"]]["revenue"] += it["price"] * it["qty"]
    top_list = sorted(top.values(), key=lambda x: x["qty"], reverse=True)[:5]

    return {
        "today_revenue": total_today,
        "week_revenue": total_week,
        "month_revenue": total_month,
        "month_profit": profit_month,
        "tx_count": len(all_tx),
        "daily_series": daily_series,
        "top_products": top_list,
    }

# --- Cashier Leaderboard (today) ---
@api.get("/analytics/cashier-leaderboard")
async def cashier_leaderboard(user: dict = Depends(get_current_user)):
    today = now_utc().date()
    cutoff = today.isoformat()
    all_tx = await db.transactions.find({"tenant_id": tenant_of(user),
                                         "status": "completed",
                                         "created_at": {"$gte": cutoff}},
                                        {"_id": 0}).to_list(5000)
    by_cashier = {}
    for t in all_tx:
        c = t.get("cashier") or "-"
        row = by_cashier.setdefault(c, {"cashier": c, "cashier_id": t.get("cashier_id"),
                                         "tx_count": 0, "revenue": 0, "profit": 0, "items_sold": 0,
                                         "cash": 0, "qris": 0})
        row["tx_count"] += 1
        row["revenue"] += t.get("total", 0)
        row["profit"] += t.get("net_profit", 0)
        row["items_sold"] += sum(i.get("qty", 0) for i in t.get("items", []))
        if t.get("payment_method") == "cash":
            row["cash"] += t.get("total", 0)
        else:
            row["qris"] += t.get("total", 0)
    rows = sorted(by_cashier.values(), key=lambda r: r["revenue"], reverse=True)
    # Attach display names (tenant-scoped users)
    users = {u["email"]: u for u in await db.users.find({"tenant_id": tenant_of(user)},
                                                        {"_id": 0, "password_hash": 0}).to_list(200)}
    for r in rows:
        r["name"] = users.get(r["cashier"], {}).get("name") or r["cashier"].split("@")[0]
    return rows

# --- Excel Import for products ---
@api.get("/products/import-template.xlsx")
async def import_template(user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    wb = Workbook(); ws = wb.active; ws.title = "Produk"
    headers = ["Nama", "Kategori", "SKU", "Stok", "Harga Modal", "Harga Jual", "URL Gambar"]
    fill = PatternFill("solid", fgColor="E85D04"); font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.fill = fill; c.font = font
    sample = [
        ["Indomie Goreng", "Mie Instan", "IDM-GRG-001", 50, 2800, 3500, ""],
        ["Kopi Kapal Api", "Minuman", "KAP-SCH", 100, 1200, 1500, ""],
    ]
    for r_idx, row in enumerate(sample, 2):
        for c_idx, v in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 22
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_produk.xlsx"'})

@api.post("/products/import")
async def import_products(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    from openpyxl import load_workbook
    from io import BytesIO
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"File tidak valid: {str(e)}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(400, "File kosong")
    header = [str(h or "").strip().lower() for h in rows[0]]
    def idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return -1
    i_name = idx("nama", "name")
    i_cat = idx("kategori", "category")
    i_sku = idx("sku", "barcode")
    i_stock = idx("stok", "stock")
    i_buy = idx("harga modal", "buy_price", "modal")
    i_sell = idx("harga jual", "sell_price", "jual")
    i_img = idx("url gambar", "image_url", "gambar")
    if i_name < 0:
        raise HTTPException(400, "Kolom 'Nama' wajib ada")
    created = updated = errors = 0
    errs = []
    for r_idx, row in enumerate(rows[1:], 2):
        try:
            name = str(row[i_name] or "").strip()
            if not name:
                continue
            payload = {
                "name": name,
                "category": str(row[i_cat] or "Umum").strip() if i_cat >= 0 else "Umum",
                "sku": str(row[i_sku] or "").strip() if i_sku >= 0 else "",
                "stock": int(row[i_stock] or 0) if i_stock >= 0 else 0,
                "buy_price": int(row[i_buy] or 0) if i_buy >= 0 else 0,
                "sell_price": int(row[i_sell] or 0) if i_sell >= 0 else 0,
                "image_url": str(row[i_img] or "").strip() if i_img >= 0 else "",
            }
            # match by SKU first, then by name (tenant-scoped)
            existing = None
            if payload["sku"]:
                existing = await db.products.find_one({"sku": payload["sku"],
                                                       "tenant_id": tenant_of(user)})
            if not existing:
                existing = await db.products.find_one({"name": {"$regex": f"^{name}$", "$options": "i"},
                                                       "tenant_id": tenant_of(user)})
            if existing:
                await db.products.update_one({"id": existing["id"], "tenant_id": tenant_of(user)},
                                             {"$set": payload})
                updated += 1
            else:
                await db.products.insert_one({"id": str(uuid.uuid4()),
                                              "tenant_id": tenant_of(user), **payload,
                                              "created_at": now_utc().isoformat()})
                created += 1
        except Exception as e:
            errors += 1
            errs.append(f"Baris {r_idx}: {str(e)}")
    return {"created": created, "updated": updated, "errors": errors, "error_details": errs[:10]}

# ============================================================
# Customers / Debt Ledger / Expenses / AI Bundling
# ============================================================

# --- Customers & Debt ---
class CustomerIn(BaseModel):
    name: str
    phone: str = ""
    note: str = ""

@api.get("/customers")
async def list_customers(user: dict = Depends(get_current_user)):
    return await db.customers.find({"tenant_id": tenant_of(user)}, {"_id": 0}).sort("total_debt", -1).to_list(500)

@api.post("/customers")
async def create_customer(body: CustomerIn, user: dict = Depends(get_current_user)):
    if body.phone:
        existing = await db.customers.find_one({"phone": body.phone,
                                                "tenant_id": tenant_of(user)}, {"_id": 0})
        if existing:
            return existing
    cid = str(uuid.uuid4())
    doc = {"id": cid, "tenant_id": tenant_of(user), **body.model_dump(),
           "total_debt": 0,
           "created_at": now_utc().isoformat()}
    await db.customers.insert_one(doc)
    return strip_id(doc)

@api.get("/customers/{cid}/ledger")
async def customer_ledger(cid: str, user: dict = Depends(get_current_user)):
    cust = await db.customers.find_one({"id": cid, "tenant_id": tenant_of(user)}, {"_id": 0})
    if not cust:
        raise HTTPException(404, "Pelanggan tidak ditemukan")
    entries = await db.debt_ledger.find({"customer_id": cid, "tenant_id": tenant_of(user)},
                                        {"_id": 0}).sort("created_at", -1).to_list(500)
    txs = await db.transactions.find({"customer_id": cid, "tenant_id": tenant_of(user)},
                                     {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"customer": cust, "entries": entries, "transactions": txs}

class DebtPayIn(BaseModel):
    amount: int = Field(gt=0)
    method: Literal["cash", "qris"] = "cash"

@api.post("/customers/{cid}/pay-debt")
async def pay_debt(cid: str, body: DebtPayIn, user: dict = Depends(get_current_user)):
    cust = await db.customers.find_one({"id": cid, "tenant_id": tenant_of(user)})
    if not cust:
        raise HTTPException(404, "Pelanggan tidak ditemukan")
    pay = min(body.amount, cust.get("total_debt", 0))
    await db.customers.update_one({"id": cid, "tenant_id": tenant_of(user)},
                                  {"$inc": {"total_debt": -pay}})
    await db.debt_ledger.insert_one({
        "id": str(uuid.uuid4()), "tenant_id": tenant_of(user),
        "customer_id": cid, "type": "credit",
        "amount": pay, "method": body.method,
        "created_at": now_utc().isoformat(),
    })
    # Mark related unpaid transactions as completed (oldest first) up to `pay` amount
    remain = pay
    unpaid = await db.transactions.find({"customer_id": cid, "tenant_id": tenant_of(user),
                                         "status": "unpaid"},
                                        {"_id": 0}).sort("created_at", 1).to_list(200)
    for t in unpaid:
        if remain <= 0:
            break
        if t["total"] <= remain:
            await db.transactions.update_one({"id": t["id"], "tenant_id": tenant_of(user)},
                                             {"$set": {"status": "completed"}})
            remain -= t["total"]
    return {"paid": pay, "remaining_debt": cust.get("total_debt", 0) - pay}

@api.post("/customers/{cid}/send-reminder")
async def send_debt_reminder(cid: str, user: dict = Depends(get_current_user)):
    cust = await db.customers.find_one({"id": cid, "tenant_id": tenant_of(user)})
    if not cust:
        raise HTTPException(404, "Pelanggan tidak ditemukan")
    debt = cust.get("total_debt", 0)
    if debt <= 0:
        raise HTTPException(400, "Tidak ada utang aktif")
    if not cust.get("phone"):
        raise HTTPException(400, "Nomor WhatsApp pelanggan kosong")
    # Try to create QRIS link
    qr_url = None
    try:
        creds = await get_midtrans_creds(tenant_of(user))
        oid = f"DEBT-{cid[:6].upper()}-{int(now_utc().timestamp())}"
        payload = {"payment_type": "qris",
                   "transaction_details": {"order_id": oid, "gross_amount": debt},
                   "qris": {"acquirer": "gopay"}}
        async with httpx.AsyncClient(timeout=20) as hc:
            r = await hc.post(f"{midtrans_base(creds['mode'])}/v2/charge",
                              json=payload, auth=(creds['server_key'], ""),
                              headers={"Accept": "application/json"})
        data = r.json()
        for a in data.get("actions", []):
            if a.get("name") in ("generate-qr-code-v2", "generate-qr-code"):
                qr_url = a["url"]; break
    except Exception:
        logging.exception("QRIS for reminder failed")
    st = await db.settings.find_one({"tenant_id": tenant_of(user)}) or {}
    store_name = st.get("store", {}).get("name", "Toko Kami")
    debt_fmt = f"Rp{debt:,}".replace(",", ".")
    msg = (f"Halo *{cust['name']}* 👋\n\n"
           f"Terima kasih telah berbelanja di *{store_name}*. "
           f"Kami ingin mengingatkan tagihan Anda saat ini sebesar *{debt_fmt}*.\n\n"
           + ("Silakan bayar langsung via QRIS pada gambar terlampir 📱.\n\n" if qr_url else "")
           + "Salam hangat,\nStaff Kami 🙏")
    try:
        await wa_send_raw(tenant_of(user), cust["phone"], msg, image_url=qr_url)
    except Exception as e:
        raise HTTPException(502, f"Gagal kirim WA: {str(e)}")
    return {"ok": True, "amount": debt, "qr_url": qr_url}

# --- Expenses ---
class ExpenseIn(BaseModel):
    name: str
    category: str = "Umum"
    amount: int = Field(gt=0)
    date: Optional[str] = None
    note: str = ""

@api.get("/expenses")
async def list_expenses(days: int = 30, user: dict = Depends(get_current_user)):
    cutoff = (now_utc() - timedelta(days=days)).isoformat()
    return await db.expenses.find({"tenant_id": tenant_of(user),
                                   "date": {"$gte": cutoff}},
                                  {"_id": 0}).sort("date", -1).to_list(500)

@api.post("/expenses")
async def create_expense(body: ExpenseIn, user: dict = Depends(get_current_user)):
    eid = str(uuid.uuid4())
    doc = {"id": eid, "tenant_id": tenant_of(user), **body.model_dump(),
           "date": body.date or now_utc().isoformat(),
           "created_at": now_utc().isoformat(),
           "created_by": user["email"]}
    await db.expenses.insert_one(doc)
    return strip_id(doc)

@api.delete("/expenses/{eid}")
async def delete_expense(eid: str, user: dict = Depends(get_current_user)):
    await db.expenses.delete_one({"id": eid, "tenant_id": tenant_of(user)})
    return {"ok": True}

@api.get("/analytics/net-profit")
async def net_profit(days: int = 30, user: dict = Depends(get_current_user)):
    cutoff = (now_utc() - timedelta(days=days)).isoformat()
    txs = await db.transactions.find({"tenant_id": tenant_of(user),
                                      "created_at": {"$gte": cutoff},
                                      "status": {"$in": ["completed", "unpaid"]}},
                                     {"_id": 0}).to_list(5000)
    revenue = sum(t["total"] for t in txs if t["status"] == "completed")
    hpp = 0  # Cost of Goods Sold
    for t in txs:
        for it in t.get("items", []):
            hpp += it.get("buy_price", 0) * it.get("qty", 0)
    gross_profit = revenue - hpp
    exp_docs = await db.expenses.find({"tenant_id": tenant_of(user),
                                       "date": {"$gte": cutoff}}, {"_id": 0}).to_list(1000)
    op_expenses = sum(e.get("amount", 0) for e in exp_docs)
    exp_by_cat = {}
    for e in exp_docs:
        exp_by_cat.setdefault(e["category"], 0)
        exp_by_cat[e["category"]] += e["amount"]
    return {
        "days": days,
        "revenue": revenue,
        "hpp": hpp,
        "gross_profit": gross_profit,
        "operational_expenses": op_expenses,
        "net_profit": gross_profit - op_expenses,
        "expenses_by_category": [{"category": k, "amount": v} for k, v in exp_by_cat.items()],
        "outstanding_debt": sum(t["total"] for t in txs if t["status"] == "unpaid"),
    }

# --- AI Dynamic Bundling ---
@api.get("/ai/suggest-bundles")
async def suggest_bundles(user: dict = Depends(get_current_user)):
    """AI detects slow-moving products and suggests bundles with popular ones."""
    try:
        # Compute movement over last 30 days
        cutoff = (now_utc() - timedelta(days=30)).isoformat()
        txs = await db.transactions.find({"tenant_id": tenant_of(user),
                                          "created_at": {"$gte": cutoff},
                                          "status": "completed"}, {"_id": 0}).to_list(5000)
        sales_by_product = {}
        for t in txs:
            for it in t.get("items", []):
                sales_by_product.setdefault(it["product_id"], 0)
                sales_by_product[it["product_id"]] += it["qty"]
        products = await db.products.find({"tenant_id": tenant_of(user)}, {"_id": 0}).to_list(500)
        # Rank
        rows = []
        for p in products:
            sold = sales_by_product.get(p["id"], 0)
            rows.append({**p, "sold_30d": sold,
                         "velocity": sold / max(1, p.get("stock", 1))})
        rows.sort(key=lambda r: r["sold_30d"])
        slow = [r for r in rows if r["sold_30d"] < 5 and r.get("stock", 0) > 0][:8]
        popular = sorted(rows, key=lambda r: r["sold_30d"], reverse=True)[:8]

        if not slow or not popular:
            return {"suggestions": [], "note": "Belum cukup data penjualan."}

        slow_str = "\n".join([f"- {p['name']} (id={p['id']}, stok={p['stock']}, terjual30h={p['sold_30d']})" for p in slow])
        pop_str = "\n".join([f"- {p['name']} (id={p['id']}, terjual30h={p['sold_30d']})" for p in popular])

        system = (
            "Kamu adalah AI merchandising untuk toko UMKM Indonesia. "
            "Diberikan daftar produk lambat & produk laris, sarankan 3 bundling menarik: "
            "pasangkan produk lambat dengan produk laris komplementer. "
            'Balas ONLY JSON dalam format persis: '
            '{"suggestions":[{"slow_id":"...","slow_name":"...","popular_id":"...",'
            '"popular_name":"...","promo_name":"...","reason":"...","discount_pct":10}]}. '
            'Jangan gunakan markdown atau code fence.'
        )
        prompt = f"PRODUK LAMBAT:\n{slow_str}\n\nPRODUK LARIS:\n{pop_str}"
        result = await gemini_generate(tenant_of(user), system, prompt)
        import re, json
        text = re.sub(r"^```(json)?", "", str(result).strip()).rstrip("`").strip()
        m = re.search(r"\{.*\}", text, re.DOTALL)
        parsed = json.loads(m.group(0)) if m else {"suggestions": []}
        return parsed
    except HTTPException:
        raise
    except Exception as e:
        logging.exception("Bundle suggest failed")
        raise HTTPException(500, f"AI bundling gagal: {str(e)}")

class ApplyBundleIn(BaseModel):
    slow_id: str
    popular_id: str
    promo_name: str
    discount_pct: int = 10

@api.post("/ai/apply-bundle")
async def apply_bundle(body: ApplyBundleIn, user: dict = Depends(get_current_user)):
    """Create a BxGy or percentage promo based on suggestion."""
    slow = await db.products.find_one({"id": body.slow_id, "tenant_id": tenant_of(user)})
    if not slow:
        raise HTTPException(404, "Produk lambat tidak ditemukan")
    pid = str(uuid.uuid4())
    # Create a percentage promo on the slow product when bought together (simplified: pct on slow when qty>=1)
    doc = {"id": pid, "tenant_id": tenant_of(user), "name": body.promo_name,
           "type": "bxgy", "product_id": body.popular_id,
           "buy_qty": 1, "get_product_id": body.slow_id, "get_qty": 1,
           "value": body.discount_pct, "min_purchase": 0,
           "active": True, "ai_generated": True,
           "created_at": now_utc().isoformat()}
    await db.promotions.insert_one(doc)
    return strip_id(doc)

# --- Update seed_data indexes ---


# --- AI: Vision OCR for wholesale receipt ---
@api.post("/ai/scan-receipt")
async def scan_receipt(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    try:
        content = await file.read()
        b64 = base64.b64encode(content).decode()
        mime = file.content_type or "image/jpeg"
        system = (
            "You are an OCR AI for Indonesian wholesale receipts (Nota Belanja Grosir). "
            "Extract each product line item and return ONLY valid JSON: "
            '{"items":[{"name":"...","qty":<int>,"buy_price":<int rupiah per unit, no decimals>}]}. '
            "Do not include markdown or explanation."
        )
        result = await gemini_generate(tenant_of(user), system,
                                       "Extract items from this receipt image.",
                                       image_b64=b64, image_mime=mime)
        import re, json
        text = re.sub(r"^```(json)?", "", str(result).strip()).rstrip("`").strip()
        m = re.search(r"\{.*\}", text, re.DOTALL)
        parsed = json.loads(m.group(0)) if m else {"items": []}
        return parsed
    except HTTPException:
        raise
    except Exception as e:
        logging.exception("OCR failed")
        raise HTTPException(500, f"AI OCR gagal: {str(e)}")

class RestockItem(BaseModel):
    product_id: Optional[str] = None
    name: str
    qty: int
    buy_price: int
    category: str = "Umum"

class RestockIn(BaseModel):
    items: List[RestockItem]

@api.post("/ai/confirm-restock")
async def confirm_restock(body: RestockIn, user: dict = Depends(get_current_user)):
    updated = 0
    created = 0
    for it in body.items:
        existing = None
        if it.product_id:
            existing = await db.products.find_one({"id": it.product_id, "tenant_id": tenant_of(user)})
        if not existing:
            existing = await db.products.find_one({"name": {"$regex": f"^{it.name}$", "$options": "i"},
                                                    "tenant_id": tenant_of(user)})
        if existing:
            await db.products.update_one({"id": existing["id"], "tenant_id": tenant_of(user)},
                                         {"$inc": {"stock": it.qty},
                                          "$set": {"buy_price": it.buy_price}})
            updated += 1
        else:
            pid = str(uuid.uuid4())
            await db.products.insert_one({
                "id": pid, "tenant_id": tenant_of(user),
                "name": it.name, "category": it.category,
                "stock": it.qty, "buy_price": it.buy_price,
                "sell_price": int(it.buy_price * 1.2),
                "sku": "", "image_url": "",
                "created_at": now_utc().isoformat()
            })
            created += 1
    return {"updated": updated, "created": created}

# --- Barcode label helpers ---
@api.post("/products/generate-sku")
async def generate_missing_sku(user: dict = Depends(get_current_user)):
    """Assign KP-XXXXXXXX SKU to products missing it."""
    prods = await db.products.find({"tenant_id": tenant_of(user),
                                    "$or": [{"sku": {"$in": [None, ""]}},
                                            {"sku": {"$exists": False}}]},
                                   {"_id": 0}).to_list(2000)
    updated = 0
    for p in prods:
        digits = "".join([c for c in p["id"] if c.isdigit()])[:8].ljust(8, "0")
        sku = f"KP{digits}"
        while await db.products.find_one({"sku": sku, "tenant_id": tenant_of(user),
                                          "id": {"$ne": p["id"]}}):
            digits = "".join([str((int(d) + 1) % 10) for d in digits])
            sku = f"KP{digits}"
        await db.products.update_one({"id": p["id"], "tenant_id": tenant_of(user)},
                                     {"$set": {"sku": sku}})
        updated += 1
    return {"updated": updated}

# --- Barcode lookup ---
@api.get("/products/by-sku/{sku}")
async def product_by_sku(sku: str, user: dict = Depends(get_current_user)):
    doc = await db.products.find_one({"sku": sku, "tenant_id": tenant_of(user)}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Produk tidak ditemukan")
    return doc

# --- WhatsApp / Fonnte ---
class FonnteSettingsIn(BaseModel):
    token: str = Field(min_length=5)

class WASendIn(BaseModel):
    target: str = Field(min_length=5)  # e.g. 628xxx
    message: str = Field(min_length=1)

class WASendReceiptIn(BaseModel):
    target: str = Field(min_length=5)
    transaction_id: str

async def get_fonnte_token(tenant_id: str) -> str:
    doc = await db.settings.find_one({"tenant_id": tenant_id}) or {}
    tk_enc = doc.get("fonnte", {}).get("token_enc")
    if not tk_enc:
        raise HTTPException(400, "Fonnte belum dikonfigurasi di Pengaturan.")
    return dec(tk_enc)

def normalize_wa_target(t: str) -> str:
    t = t.strip().replace("+", "").replace("-", "").replace(" ", "")
    if t.startswith("0"):
        t = "62" + t[1:]
    return t

@api.put("/settings/fonnte")
async def save_fonnte(body: FonnteSettingsIn, user: dict = Depends(get_current_user)):
    await db.settings.update_one({"tenant_id": tenant_of(user)},
                                 {"$set": {"tenant_id": tenant_of(user), "fonnte": {
                                     "token_enc": enc(body.token),
                                     "updated_at": now_utc().isoformat(),
                                 }}}, upsert=True)
    return {"ok": True}

@api.post("/whatsapp/send")
async def wa_send(body: WASendIn, user: dict = Depends(get_current_user)):
    token = await get_fonnte_token(tenant_of(user))
    target = normalize_wa_target(body.target)
    try:
        async with httpx.AsyncClient(timeout=20) as hc:
            r = await hc.post("https://api.fonnte.com/send",
                              headers={"Authorization": token},
                              data={"target": target, "message": body.message,
                                    "countryCode": "62"})
        data = r.json() if r.headers.get("content-type", "").startswith("application/json") else {"raw": r.text}
    except Exception as e:
        raise HTTPException(502, f"Gagal terhubung ke Fonnte: {str(e)}")
    if not data.get("status"):
        raise HTTPException(400, data.get("reason") or "Gagal mengirim WhatsApp")
    return {"ok": True, "detail": data.get("detail"), "requestid": data.get("requestid")}

def format_receipt_wa(store: dict, tx: dict) -> str:
    def fmt(n): return "Rp" + f"{int(n or 0):,}".replace(",", ".")
    lines = []
    lines.append(f"*{store.get('name', 'Toko')}*")
    if store.get("address"): lines.append(store["address"])
    if store.get("phone"): lines.append(store["phone"])
    lines.append("--------------------------------")
    lines.append(f"No : {tx['order_id']}")
    lines.append(f"Tgl: {datetime.fromisoformat(tx['created_at']).strftime('%d/%m/%Y %H:%M')}")
    lines.append("--------------------------------")
    for it in tx["items"]:
        lines.append(f"{it['name']}")
        lines.append(f"  {it['qty']} x {fmt(it['price'])} = {fmt(it['qty'] * it['price'])}")
    lines.append("--------------------------------")
    lines.append(f"Subtotal : {fmt(tx['subtotal'])}")
    if tx.get("discount"): lines.append(f"Diskon   : -{fmt(tx['discount'])}")
    if tx.get("tax"):      lines.append(f"Pajak    : {fmt(tx['tax'])}")
    lines.append(f"*TOTAL    : {fmt(tx['total'])}*")
    lines.append(f"Bayar ({tx.get('payment_method', '').upper()}): {fmt(tx.get('cash_tendered') or tx['total'])}")
    if tx.get("change"): lines.append(f"Kembalian: {fmt(tx['change'])}")
    lines.append("--------------------------------")
    lines.append(store.get("footer", "Terima Kasih!"))
    return "\n".join(lines)

@api.post("/whatsapp/send-receipt")
async def wa_send_receipt(body: WASendReceiptIn, user: dict = Depends(get_current_user)):
    tx = await db.transactions.find_one({"id": body.transaction_id,
                                         "tenant_id": tenant_of(user)}, {"_id": 0})
    if not tx:
        raise HTTPException(404, "Transaksi tidak ditemukan")
    st = await db.settings.find_one({"tenant_id": tenant_of(user)}) or {}
    store = st.get("store", {})
    msg = format_receipt_wa(store, tx)
    return await wa_send(WASendIn(target=body.target, message=msg), user)

# --- AI Order Parser (parses free-text WA message into cart items) ---
class OrderParseIn(BaseModel):
    text: str = Field(min_length=1)

@api.post("/ai/parse-order")
async def parse_order(body: OrderParseIn, user: dict = Depends(get_current_user)):
    try:
        products = await db.products.find({"tenant_id": tenant_of(user)}, {"_id": 0}).to_list(500)
        catalog = "\n".join([f"- {p['name']} (id={p['id']}, harga={p['sell_price']}, stok={p['stock']})" for p in products])

        system = (
            "Kamu adalah AI yang memahami pesan pemesanan WhatsApp dari pelanggan toko UMKM Indonesia. "
            "Diberikan katalog produk dan pesan bebas dari pelanggan, cocokkan item yang dipesan "
            "berdasarkan nama yang mirip (misal 'indomie' cocok dengan 'Indomie Goreng'). "
            "Balas ONLY JSON valid dalam format: "
            '{"items":[{"product_id":"...","name":"...","qty":<int>}], "unmatched":["..."]}. '
            "unmatched berisi item yang tidak ditemukan di katalog. "
            "Jangan gunakan markdown."
        )
        prompt = f"KATALOG:\n{catalog}\n\nPESAN PELANGGAN:\n{body.text}"
        result = await gemini_generate(tenant_of(user), system, prompt)
        import re, json
        text = re.sub(r"^```(json)?", "", str(result).strip()).rstrip("`").strip()
        m = re.search(r"\{.*\}", text, re.DOTALL)
        parsed = json.loads(m.group(0)) if m else {"items": [], "unmatched": []}

        # Enrich with current price & stock
        by_id = {p["id"]: p for p in products}
        enriched = []
        for it in parsed.get("items", []):
            pr = by_id.get(it.get("product_id"))
            if pr:
                enriched.append({
                    "product_id": pr["id"], "name": pr["name"],
                    "qty": int(it.get("qty", 1)),
                    "price": pr["sell_price"], "buy_price": pr["buy_price"],
                    "stock": pr["stock"],
                })
        return {"items": enriched, "unmatched": parsed.get("unmatched", [])}
    except HTTPException:
        raise
    except Exception as e:
        logging.exception("Order parse failed")
        raise HTTPException(500, f"AI gagal memahami pesan: {str(e)}")


# --- AI Business Insights ---
@api.get("/ai/insights")
async def ai_insights(user: dict = Depends(get_current_user)):
    try:
        summary = await analytics_summary(user)
        products = await db.products.find({"tenant_id": tenant_of(user)}, {"_id": 0}).to_list(500)
        low_stock = [p["name"] for p in products if p["stock"] <= 5]
        top_names = ", ".join([f"{p['name']} ({p['qty']} terjual)" for p in summary["top_products"][:3]]) or "belum ada data"

        system = (
            "Kamu adalah asisten AI bisnis untuk pemilik toko UMKM Indonesia. "
            "Berikan 3-4 kalimat singkat, ramah, dalam Bahasa Indonesia. "
            "Fokus pada: produk terlaris, peringatan stok menipis, dan saran singkat. "
            "Jangan pakai markdown atau bullet. Tulis sebagai paragraf natural."
        )
        prompt = (
            f"Pendapatan hari ini: Rp{summary['today_revenue']:,}. "
            f"Pendapatan 30 hari: Rp{summary['month_revenue']:,}. "
            f"Laba bersih 30 hari: Rp{summary['month_profit']:,}. "
            f"Total transaksi: {summary['tx_count']}. "
            f"Produk terlaris: {top_names}. "
            f"Stok menipis: {', '.join(low_stock[:5]) or 'tidak ada'}."
        )
        result = await gemini_generate(tenant_of(user), system, prompt)
        return {"insight": str(result).strip()}
    except HTTPException as he:
        # Surface the 400 to frontend so it can prompt to configure the key
        if he.status_code == 400:
            raise
        return {"insight": f"Ringkasan AI tidak tersedia saat ini. ({he.detail})"}
    except Exception as e:
        logging.exception("Insight failed")
        return {"insight": f"Ringkasan AI tidak tersedia saat ini. ({str(e)})"}

# --- Settings (Store + Midtrans) ---
class StoreProfile(BaseModel):
    name: str = "Toko Saya"
    address: str = ""
    phone: str = ""
    footer: str = "Terima Kasih atas Kunjungan Anda!"

class MidtransSettingsIn(BaseModel):
    mode: Literal["sandbox", "production"] = "sandbox"
    merchant_id: str = ""
    client_key: str = ""
    server_key: str = ""

@api.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    doc = await db.settings.find_one({"tenant_id": tenant_of(user)}, {"_id": 0}) or {}
    store = doc.get("store", StoreProfile().model_dump())
    mt = doc.get("midtrans", {})
    fn = doc.get("fonnte", {})
    gm = doc.get("gemini", {})
    br = doc.get("branding", {})
    return {
        "store": store,
        "midtrans": {
            "mode": mt.get("mode", "sandbox"),
            "merchant_id": dec(mt.get("merchant_id_enc", "")),
            "client_key": dec(mt.get("client_key_enc", "")),
            "server_key_masked": ("*" * 8 + dec(mt.get("server_key_enc", ""))[-4:]) if mt.get("server_key_enc") else "",
            "configured": bool(mt.get("server_key_enc")),
        },
        "fonnte": {
            "configured": bool(fn.get("token_enc")),
        },
        "gemini": {
            "configured": bool(gm.get("api_key_enc")),
            "model": GEMINI_MODEL_TEXT,
        },
        "branding": {
            "app_name": br.get("app_name") or "KasirPintar AI",
            "short_name": br.get("short_name") or "KasirPintar",
            "theme_color": br.get("theme_color") or "#e85d04",
            "logo_base64": br.get("logo_base64") or "",
            "updated_at": br.get("updated_at"),
        },
    }

@api.put("/settings/store")
async def save_store(body: StoreProfile, user: dict = Depends(get_current_user)):
    await db.settings.update_one({"tenant_id": tenant_of(user)},
                                 {"$set": {"tenant_id": tenant_of(user),
                                           "store": body.model_dump()}},
                                 upsert=True)
    return body

@api.put("/settings/midtrans")
async def save_midtrans(body: MidtransSettingsIn, user: dict = Depends(get_current_user)):
    await db.settings.update_one({"tenant_id": tenant_of(user)},
                                 {"$set": {"tenant_id": tenant_of(user), "midtrans": {
                                     "mode": body.mode,
                                     "merchant_id_enc": enc(body.merchant_id),
                                     "client_key_enc": enc(body.client_key),
                                     "server_key_enc": enc(body.server_key),
                                     "updated_at": now_utc().isoformat(),
                                 }}}, upsert=True)
    return {"ok": True}

async def get_midtrans_creds(tenant_id: str):
    doc = await db.settings.find_one({"tenant_id": tenant_id}) or {}
    mt = doc.get("midtrans", {})
    if not mt.get("server_key_enc"):
        raise HTTPException(400, "Midtrans belum dikonfigurasi. Silakan atur di Pengaturan.")
    return {"mode": mt.get("mode", "sandbox"),
            "server_key": dec(mt["server_key_enc"]),
            "client_key": dec(mt.get("client_key_enc", "")),
            "merchant_id": dec(mt.get("merchant_id_enc", ""))}


# --- Gemini AI settings (BYOK per-tenant) ---
class GeminiSettingsIn(BaseModel):
    api_key: str = Field(min_length=8)

@api.put("/settings/gemini")
async def save_gemini(body: GeminiSettingsIn, user: dict = Depends(get_current_user)):
    await db.settings.update_one({"tenant_id": tenant_of(user)},
                                 {"$set": {"tenant_id": tenant_of(user), "gemini": {
                                     "api_key_enc": enc(body.api_key.strip()),
                                     "updated_at": now_utc().isoformat(),
                                 }}}, upsert=True)
    return {"ok": True, "configured": True}

@api.delete("/settings/gemini")
async def clear_gemini(user: dict = Depends(get_current_user)):
    await db.settings.update_one({"tenant_id": tenant_of(user)},
                                 {"$unset": {"gemini": ""}})
    return {"ok": True, "configured": False}

@api.post("/settings/gemini/test")
async def test_gemini(user: dict = Depends(get_current_user)):
    """Quick sanity check: prompt Gemini with a trivial ping."""
    try:
        out = await gemini_generate(tenant_of(user),
                                    "Balas persis kata: OK",
                                    "ping")
        return {"ok": True, "reply": (out or "").strip()[:120]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(502, f"Test gagal: {str(e)}")


# --- Per-tenant Branding (white-label + dynamic PWA manifest) ---
import re as _re_branding

_HEX_RE = _re_branding.compile(r"^#[0-9a-fA-F]{6}$")

class BrandingIn(BaseModel):
    app_name: str = Field(min_length=1, max_length=30)
    short_name: str = Field(min_length=1, max_length=12)
    theme_color: str = "#e85d04"
    logo_base64: str = ""  # data URL 'data:image/...;base64,...' or empty

@api.get("/settings/branding")
async def get_branding(user: dict = Depends(get_current_user)):
    doc = await db.settings.find_one({"tenant_id": tenant_of(user)}) or {}
    b = doc.get("branding") or {}
    return {
        "app_name": b.get("app_name") or "KasirPintar AI",
        "short_name": b.get("short_name") or "KasirPintar",
        "theme_color": b.get("theme_color") or "#e85d04",
        "logo_base64": b.get("logo_base64") or "",
        "updated_at": b.get("updated_at"),
    }

@api.put("/settings/branding")
async def save_branding(body: BrandingIn, user: dict = Depends(get_current_user)):
    color = body.theme_color.strip()
    if not _HEX_RE.match(color):
        raise HTTPException(400, "Warna tema harus format hex #RRGGBB (misal #e85d04)")
    logo = (body.logo_base64 or "").strip()
    # Optional server-side re-crop/resize to 512x512 (best-effort)
    if logo.startswith("data:image"):
        try:
            from PIL import Image
            import io
            header, b64part = logo.split(",", 1)
            img_bytes = base64.b64decode(b64part)
            img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
            w, h = img.size
            side = min(w, h)
            left = (w - side) // 2
            top = (h - side) // 2
            img = img.crop((left, top, left + side, top + side)).resize((512, 512), Image.LANCZOS)
            buf = io.BytesIO()
            img.save(buf, format="PNG", optimize=True)
            logo = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
        except Exception:
            logging.exception("Logo resize failed; storing client-provided image as-is")
    if logo and len(logo) > 800_000:
        raise HTTPException(400, "Logo terlalu besar. Maksimum 600KB setelah encoding.")
    payload = {
        "app_name": body.app_name.strip(),
        "short_name": body.short_name.strip(),
        "theme_color": color,
        "logo_base64": logo,
        "updated_at": now_utc().isoformat(),
    }
    await db.settings.update_one(
        {"tenant_id": tenant_of(user)},
        {"$set": {"tenant_id": tenant_of(user), "branding": payload}},
        upsert=True,
    )
    return {"ok": True, **payload}

@api.delete("/settings/branding")
async def reset_branding(user: dict = Depends(get_current_user)):
    await db.settings.update_one({"tenant_id": tenant_of(user)}, {"$unset": {"branding": ""}})
    return {"ok": True}

def midtrans_base(mode: str) -> str:
    return "https://api.sandbox.midtrans.com" if mode == "sandbox" else "https://api.midtrans.com"

# --- Midtrans QRIS ---
class QRISCreateIn(BaseModel):
    order_id: str
    amount: int = Field(gt=0)

@api.post("/payments/qris")
async def create_qris(body: QRISCreateIn, user: dict = Depends(get_current_user)):
    creds = await get_midtrans_creds(tenant_of(user))
    payload = {
        "payment_type": "qris",
        "transaction_details": {"order_id": body.order_id, "gross_amount": body.amount},
        "qris": {"acquirer": "gopay"}
    }
    async with httpx.AsyncClient(timeout=20) as hc:
        r = await hc.post(f"{midtrans_base(creds['mode'])}/v2/charge",
                          json=payload, auth=(creds['server_key'], ""),
                          headers={"Accept": "application/json"})
    data = r.json()
    if r.status_code >= 400:
        raise HTTPException(r.status_code, detail=data)
    action = None
    for a in data.get("actions", []):
        if a.get("name") == "generate-qr-code-v2":
            action = a
            break
    if not action:
        for a in data.get("actions", []):
            if a.get("name") == "generate-qr-code":
                action = a
                break
    await db.qris_payments.insert_one({
        "tenant_id": tenant_of(user),
        "order_id": body.order_id, "amount": body.amount,
        "transaction_id": data.get("transaction_id"),
        "status": data.get("transaction_status", "pending"),
        "qr_url": action.get("url") if action else None,
        "created_at": now_utc().isoformat()
    })
    return {"order_id": body.order_id, "status": data.get("transaction_status"),
            "qr_url": action.get("url") if action else None,
            "transaction_id": data.get("transaction_id")}

@api.get("/payments/qris/{order_id}/status")
async def qris_status(order_id: str, user: dict = Depends(get_current_user)):
    creds = await get_midtrans_creds(tenant_of(user))
    async with httpx.AsyncClient(timeout=15) as hc:
        r = await hc.get(f"{midtrans_base(creds['mode'])}/v2/{order_id}/status",
                         auth=(creds['server_key'], ""),
                         headers={"Accept": "application/json"})
    if r.status_code >= 400:
        raise HTTPException(r.status_code, detail=r.json())
    data = r.json()
    status = data.get("transaction_status")
    await db.qris_payments.update_one({"order_id": order_id, "tenant_id": tenant_of(user)},
                                      {"$set": {"status": status,
                                                "updated_at": now_utc().isoformat()}})
    return {"order_id": order_id, "status": status,
            "fraud_status": data.get("fraud_status")}

@api.post("/payments/midtrans/webhook")
async def midtrans_webhook(request: Request):
    payload = await request.json()
    order_id = str(payload.get("order_id", ""))
    status_code = str(payload.get("status_code", ""))
    gross_amount = str(payload.get("gross_amount", ""))
    supplied = str(payload.get("signature_key", ""))
    if not order_id or not supplied:
        raise HTTPException(400, "invalid notification")
    # Look up tenant via qris_payments or online_orders
    qp = await db.qris_payments.find_one({"order_id": order_id})
    oo = await db.online_orders.find_one({"midtrans_order_id": order_id})
    tenant_id = (qp or {}).get("tenant_id") or (oo or {}).get("tenant_id")
    if not tenant_id:
        raise HTTPException(404, "Order tidak ditemukan")
    creds = await get_midtrans_creds(tenant_id)
    expected = hashlib.sha512(f"{order_id}{status_code}{gross_amount}{creds['server_key']}".encode()).hexdigest()
    if not hmac.compare_digest(expected, supplied):
        raise HTTPException(403, "invalid signature")
    tx_status = payload.get("transaction_status")
    await db.qris_payments.update_one({"order_id": order_id, "tenant_id": tenant_id},
                                      {"$set": {"status": tx_status,
                                                "last_webhook": payload,
                                                "updated_at": now_utc().isoformat()}},
                                      upsert=True)
    # If it's a WA-initiated order, notify customer & advance state
    if tx_status in ("settlement", "capture"):
        wa_order = await db.online_orders.find_one({"midtrans_order_id": order_id,
                                                    "tenant_id": tenant_id})
        if wa_order and wa_order.get("status") != "paid":
            await db.online_orders.update_one({"id": wa_order["id"], "tenant_id": tenant_id},
                {"$set": {"status": "paid", "paid_at": now_utc().isoformat()}})
            await db.wa_sessions.update_one({"phone": wa_order["customer_phone"],
                                              "tenant_id": tenant_id},
                {"$set": {"state": "done"}})
            try:
                await wa_send_raw(tenant_id, wa_order["customer_phone"],
                    f"✅ *Pembayaran Berhasil!* Pesanan #{wa_order['id'][:6].upper()} sedang diproses. Terima kasih 🙏")
            except Exception:
                pass
    return {"ok": True}


# ============================================================
# Multi-Cashier & Shifts, Promotions, Online Orders,
# Automated WhatsApp Bot, Low-Stock Alerts, Excel Exports
# ============================================================

# --- Users (cashier management) ---
class CashierIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = "Kasir"

@api.get("/users")
async def list_users(user: dict = Depends(get_current_user)):
    users = await db.users.find({"tenant_id": tenant_of(user)},
                                {"_id": 0, "password_hash": 0}).to_list(200)
    return users

@api.post("/users/cashier")
async def create_cashier(body: CashierIn, user: dict = Depends(get_current_user)):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email sudah terdaftar")
    uid = str(uuid.uuid4())
    await db.users.insert_one({
        "id": uid, "email": email, "name": body.name, "role": "cashier",
        "tenant_id": tenant_of(user),  # inherit tenant from owner
        "password_hash": hash_password(body.password),
        "created_at": now_utc().isoformat(),
    })
    return {"id": uid, "email": email, "name": body.name, "role": "cashier"}

@api.delete("/users/{uid}")
async def delete_cashier(uid: str, user: dict = Depends(get_current_user)):
    doc = await db.users.find_one({"id": uid, "tenant_id": tenant_of(user)})
    if not doc:
        raise HTTPException(404, "Kasir tidak ditemukan")
    if doc.get("role") == "owner":
        raise HTTPException(400, "Tidak bisa menghapus owner")
    await db.users.delete_one({"id": uid, "tenant_id": tenant_of(user)})
    return {"ok": True}

# --- Shifts ---
class ShiftOpenIn(BaseModel):
    opening_cash: int = Field(ge=0)

class ShiftCloseIn(BaseModel):
    closing_cash_actual: int = Field(ge=0)
    notes: str = ""

async def get_open_shift(user_id: str, tenant_id: str) -> Optional[dict]:
    return await db.shifts.find_one({"user_id": user_id, "tenant_id": tenant_id,
                                     "status": "open"}, {"_id": 0})

async def compute_shift_totals(shift_id: str, tenant_id: str) -> dict:
    txs = await db.transactions.find({"shift_id": shift_id, "tenant_id": tenant_id,
                                      "status": "completed"},
                                     {"_id": 0}).to_list(2000)
    cash_total = sum(t["total"] for t in txs if t["payment_method"] == "cash")
    qris_total = sum(t["total"] for t in txs if t["payment_method"] == "qris")
    return {"tx_count": len(txs), "cash_total": cash_total, "qris_total": qris_total,
            "grand_total": cash_total + qris_total,
            "profit": sum(t.get("net_profit", 0) for t in txs)}

@api.get("/shifts/current")
async def shift_current(user: dict = Depends(get_current_user)):
    s = await get_open_shift(user["id"], tenant_of(user))
    if not s:
        return {"open": False}
    totals = await compute_shift_totals(s["id"], tenant_of(user))
    return {"open": True, "shift": s, "totals": totals}

@api.post("/shifts/open")
async def shift_open(body: ShiftOpenIn, user: dict = Depends(get_current_user)):
    existing = await get_open_shift(user["id"], tenant_of(user))
    if existing:
        raise HTTPException(400, "Shift sudah dibuka. Tutup dulu.")
    sid = str(uuid.uuid4())
    doc = {"id": sid, "tenant_id": tenant_of(user),
           "user_id": user["id"], "user_email": user["email"],
           "user_name": user.get("name", ""),
           "opening_cash": body.opening_cash, "status": "open",
           "opened_at": now_utc().isoformat()}
    await db.shifts.insert_one(doc)
    return strip_id(doc)

@api.post("/shifts/close")
async def shift_close(body: ShiftCloseIn, user: dict = Depends(get_current_user)):
    s = await get_open_shift(user["id"], tenant_of(user))
    if not s:
        raise HTTPException(400, "Tidak ada shift aktif")
    totals = await compute_shift_totals(s["id"], tenant_of(user))
    expected_cash = s["opening_cash"] + totals["cash_total"]
    discrepancy = body.closing_cash_actual - expected_cash
    await db.shifts.update_one({"id": s["id"], "tenant_id": tenant_of(user)}, {"$set": {
        "status": "closed", "closed_at": now_utc().isoformat(),
        "closing_cash_expected": expected_cash,
        "closing_cash_actual": body.closing_cash_actual,
        "discrepancy": discrepancy, "notes": body.notes,
        "totals": totals,
    }})
    doc = await db.shifts.find_one({"id": s["id"], "tenant_id": tenant_of(user)}, {"_id": 0})
    return doc

@api.get("/shifts")
async def list_shifts(user: dict = Depends(get_current_user)):
    items = await db.shifts.find({"tenant_id": tenant_of(user)},
                                 {"_id": 0}).sort("opened_at", -1).to_list(100)
    return items

# --- Promotions ---
class PromotionIn(BaseModel):
    name: str
    type: Literal["percentage", "fixed", "bxgy", "min_purchase"]
    value: int = 0          # percentage or fixed rp
    product_id: Optional[str] = None   # for pct/fixed on specific product
    buy_qty: int = 0        # for bxgy
    get_product_id: Optional[str] = None
    get_qty: int = 0
    min_purchase: int = 0   # for min_purchase / free-delivery
    active: bool = True

@api.get("/promotions")
async def list_promos(user: dict = Depends(get_current_user)):
    return await db.promotions.find({"tenant_id": tenant_of(user)},
                                    {"_id": 0}).sort("name", 1).to_list(200)

@api.post("/promotions")
async def create_promo(body: PromotionIn, user: dict = Depends(get_current_user)):
    pid = str(uuid.uuid4())
    doc = {"id": pid, "tenant_id": tenant_of(user),
           **body.model_dump(), "created_at": now_utc().isoformat()}
    await db.promotions.insert_one(doc)
    return strip_id(doc)

@api.put("/promotions/{pid}")
async def update_promo(pid: str, body: PromotionIn, user: dict = Depends(get_current_user)):
    await db.promotions.update_one({"id": pid, "tenant_id": tenant_of(user)},
                                   {"$set": body.model_dump()})
    return await db.promotions.find_one({"id": pid, "tenant_id": tenant_of(user)}, {"_id": 0})

@api.delete("/promotions/{pid}")
async def delete_promo(pid: str, user: dict = Depends(get_current_user)):
    await db.promotions.delete_one({"id": pid, "tenant_id": tenant_of(user)})
    return {"ok": True}

@api.post("/promotions/preview")
async def promo_preview(payload: dict, user: dict = Depends(get_current_user)):
    """Given cart items [{product_id, name, qty, price}], return applied promos + total discount."""
    items = payload.get("items", [])
    subtotal = sum(i["price"] * i["qty"] for i in items)
    result = await apply_promotions(tenant_of(user), items, subtotal)
    return result

async def apply_promotions(tenant_id: str, items: List[dict], subtotal: int) -> dict:
    promos = await db.promotions.find({"tenant_id": tenant_id, "active": True},
                                      {"_id": 0}).to_list(200)
    applied = []
    total_discount = 0
    free_items = []
    free_delivery = False
    for p in promos:
        if p["type"] == "percentage":
            if p.get("product_id"):
                for it in items:
                    if it["product_id"] == p["product_id"]:
                        d = int(it["price"] * it["qty"] * (p["value"] / 100))
                        total_discount += d
                        applied.append({"name": p["name"], "discount": d})
            else:
                d = int(subtotal * (p["value"] / 100))
                total_discount += d
                applied.append({"name": p["name"], "discount": d})
        elif p["type"] == "fixed":
            if p.get("product_id"):
                for it in items:
                    if it["product_id"] == p["product_id"]:
                        d = min(it["price"] * it["qty"], p["value"])
                        total_discount += d
                        applied.append({"name": p["name"], "discount": d})
            else:
                d = min(subtotal, p["value"])
                total_discount += d
                applied.append({"name": p["name"], "discount": d})
        elif p["type"] == "bxgy":
            for it in items:
                if it["product_id"] == p.get("product_id") and it["qty"] >= p["buy_qty"] > 0:
                    times = it["qty"] // p["buy_qty"]
                    free_items.append({"product_id": p.get("get_product_id"),
                                       "qty": p["get_qty"] * times,
                                       "promo": p["name"]})
                    applied.append({"name": p["name"], "discount": 0, "free": p["get_qty"] * times})
        elif p["type"] == "min_purchase":
            if subtotal >= p["min_purchase"] and p["min_purchase"] > 0:
                if p["value"] > 0:
                    total_discount += p["value"]
                    applied.append({"name": p["name"], "discount": p["value"]})
                else:
                    free_delivery = True
                    applied.append({"name": p["name"], "free_delivery": True})
    return {"applied": applied, "total_discount": total_discount,
            "free_items": free_items, "free_delivery": free_delivery,
            "final_total": max(0, subtotal - total_discount)}

# --- Low-stock alert helper ---
async def check_low_stock_and_alert(tenant_id: str, product_ids: List[str]):
    """After stock decrement, notify owner via WA if any product is at/under threshold (<=5)."""
    st = await db.settings.find_one({"tenant_id": tenant_id}) or {}
    fn = st.get("fonnte", {})
    store = st.get("store", {})
    if not fn.get("token_enc") or not store.get("phone"):
        return
    triggered = []
    for pid in set(product_ids):
        p = await db.products.find_one({"id": pid, "tenant_id": tenant_id}, {"_id": 0})
        if not p:
            continue
        threshold = p.get("min_stock", 5) or 5
        # only alert on transition into low (avoid spam) — use last_alert timestamp check
        if p["stock"] <= threshold:
            last = p.get("last_low_alert_at")
            recent = last and (datetime.fromisoformat(last) > now_utc() - timedelta(hours=6))
            if not recent:
                triggered.append(p)
                await db.products.update_one({"id": pid, "tenant_id": tenant_id},
                                             {"$set": {"last_low_alert_at": now_utc().isoformat()}})
    if not triggered:
        return
    try:
        token = dec(fn["token_enc"])
        msg = "⚠️ *PERINGATAN STOK*\n\n"
        for p in triggered:
            msg += f"• {p['name']} sisa *{p['stock']}* pcs\n"
        msg += "\nSegera lakukan restok!"
        target = normalize_wa_target(store["phone"])
        async with httpx.AsyncClient(timeout=15) as hc:
            await hc.post("https://api.fonnte.com/send",
                          headers={"Authorization": token},
                          data={"target": target, "message": msg, "countryCode": "62"})
    except Exception:
        logging.exception("Low-stock alert failed")

# --- WA send helpers (no auth, for bot) ---
async def wa_send_raw(tenant_id: str, target: str, message: str, image_url: Optional[str] = None) -> dict:
    st = await db.settings.find_one({"tenant_id": tenant_id}) or {}
    tk_enc = st.get("fonnte", {}).get("token_enc")
    if not tk_enc:
        raise RuntimeError("Fonnte not configured")
    token = dec(tk_enc)
    data = {"target": normalize_wa_target(target), "message": message, "countryCode": "62"}
    if image_url:
        data["url"] = image_url
    async with httpx.AsyncClient(timeout=25) as hc:
        r = await hc.post("https://api.fonnte.com/send",
                          headers={"Authorization": token}, data=data)
    try:
        return r.json()
    except Exception:
        return {"raw": r.text}

# --- Online Orders ---
@api.get("/online-orders")
async def list_online_orders(user: dict = Depends(get_current_user)):
    return await db.online_orders.find({"tenant_id": tenant_of(user)},
                                       {"_id": 0}).sort("created_at", -1).to_list(200)

@api.post("/online-orders/{oid}/mark-shipped")
async def mark_shipped(oid: str, user: dict = Depends(get_current_user)):
    doc = await db.online_orders.find_one({"id": oid, "tenant_id": tenant_of(user)})
    if not doc:
        raise HTTPException(404, "Order tidak ditemukan")
    if doc.get("status") == "shipped":
        return {"ok": True, "note": "sudah dikirim"}
    # Decrement stock only when shipped (was reserved on payment)
    for it in doc.get("items", []):
        await db.products.update_one({"id": it["product_id"], "tenant_id": tenant_of(user)},
                                     {"$inc": {"stock": -it["qty"]}})
    await db.online_orders.update_one({"id": oid, "tenant_id": tenant_of(user)},
                                      {"$set": {"status": "shipped",
                                                "shipped_at": now_utc().isoformat(),
                                                "shipped_by": user["email"]}})
    # Notify customer
    try:
        await wa_send_raw(tenant_of(user), doc["customer_phone"],
                          f"📦 Pesanan #{doc['id'][:6].upper()} telah dikirim ke {doc.get('address', 'alamat Anda')}. Terima kasih!")
    except Exception:
        pass
    await check_low_stock_and_alert(tenant_of(user), [it["product_id"] for it in doc.get("items", [])])
    return await db.online_orders.find_one({"id": oid, "tenant_id": tenant_of(user)}, {"_id": 0})

# --- WhatsApp bot state machine ---
# States: idle -> awaiting_confirm -> awaiting_address -> awaiting_payment -> done
async def bot_reply(tenant_id: str, phone: str, incoming: str):
    """Process incoming WA message and reply. All persistence via db.wa_sessions."""
    sess = await db.wa_sessions.find_one({"phone": phone, "tenant_id": tenant_id})
    if not sess:
        sess = {"phone": phone, "tenant_id": tenant_id, "state": "idle",
                "cart": [], "created_at": now_utc().isoformat()}
        await db.wa_sessions.insert_one(sess)

    txt = (incoming or "").strip()
    lower = txt.lower()

    # Global commands
    if lower in ("batal", "cancel", "reset"):
        await db.wa_sessions.update_one({"phone": phone, "tenant_id": tenant_id},
                                        {"$set": {"state": "idle", "cart": []}})
        await wa_send_raw(tenant_id, phone, "Pesanan dibatalkan. Ketik menu produk yang ingin dipesan untuk mulai lagi.")
        return

    state = sess.get("state", "idle")

    if state in ("idle", "done"):
        # Parse order via AI (per-tenant Gemini key). If not configured, skip AI gracefully.
        parsed = {"items": [], "unmatched": []}
        products = await db.products.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(500)
        try:
            catalog = "\n".join([f"- {p['name']} (id={p['id']}, harga={p['sell_price']}, stok={p['stock']})" for p in products])
            system = ("Cocokkan item pesanan pelanggan dengan katalog. Balas ONLY JSON: "
                      '{"items":[{"product_id":"...","name":"...","qty":<int>}],"unmatched":["..."]}.'
                      "Jangan pakai markdown.")
            r = await gemini_generate(tenant_id, system, f"KATALOG:\n{catalog}\n\nPESAN:\n{txt}")
            import re, json
            t = re.sub(r"^```(json)?", "", str(r).strip()).rstrip("`").strip()
            m = re.search(r"\{.*\}", t, re.DOTALL)
            parsed = json.loads(m.group(0)) if m else {"items": [], "unmatched": []}
        except Exception:
            logging.exception("WA bot AI parse skipped")

        by_id = {p["id"]: p for p in products}
        cart = []
        for it in parsed.get("items", []):
            pr = by_id.get(it.get("product_id"))
            if pr and pr["stock"] >= int(it.get("qty", 1)):
                cart.append({"product_id": pr["id"], "name": pr["name"],
                             "qty": int(it["qty"]), "price": pr["sell_price"],
                             "buy_price": pr["buy_price"]})
        if not cart:
            await wa_send_raw(tenant_id, phone,
                "Halo! 👋 Sepertinya saya belum menemukan produk yang cocok. "
                "Coba sebutkan nama produk lebih spesifik, contoh:\n"
                "- Indomie Goreng 2\n- Kopi Kapal Api 3")
            return

        subtotal = sum(c["price"] * c["qty"] for c in cart)
        promo = await apply_promotions(tenant_id, cart, subtotal)
        total = promo["final_total"]

        summary = "\n".join([f"- {c['qty']}× {c['name']} = Rp{c['price']*c['qty']:,}".replace(",", ".") for c in cart])
        promo_lines = "\n".join([f"🎁 {a['name']}: -Rp{a.get('discount',0):,}".replace(",", ".") for a in promo["applied"]])
        msg = (f"Pesanan Anda:\n{summary}\n"
               f"{promo_lines + chr(10) if promo_lines else ''}"
               f"*Total: Rp{total:,}*\n\n"
               "Ketik *YA* untuk lanjut, atau kirim ulang pesanan.").replace(",", ".")
        await db.wa_sessions.update_one({"phone": phone, "tenant_id": tenant_id},
            {"$set": {"state": "awaiting_confirm", "cart": cart, "total": total,
                      "subtotal": subtotal, "promo": promo,
                      "updated_at": now_utc().isoformat()}})
        await wa_send_raw(tenant_id, phone, msg)
        return

    if state == "awaiting_confirm":
        if lower in ("ya", "yes", "y", "ok", "oke"):
            await db.wa_sessions.update_one({"phone": phone, "tenant_id": tenant_id},
                {"$set": {"state": "awaiting_name"}})
            await wa_send_raw(tenant_id, phone, "Baik! Boleh saya minta *nama penerima*?")
            return
        # else re-parse
        await db.wa_sessions.update_one({"phone": phone, "tenant_id": tenant_id},
                                        {"$set": {"state": "idle"}})
        return await bot_reply(tenant_id, phone, txt)

    if state == "awaiting_name":
        await db.wa_sessions.update_one({"phone": phone, "tenant_id": tenant_id},
            {"$set": {"state": "awaiting_address", "customer_name": txt}})
        await wa_send_raw(tenant_id, phone, "Terima kasih 🙏 Mohon kirim *alamat lengkap pengiriman* Anda.")
        return

    if state == "awaiting_address":
        # Create online order and QRIS
        cart = sess.get("cart", [])
        total = sess.get("total", 0)
        oid = str(uuid.uuid4())
        order_id_mt = f"WA{oid[:10].upper()}"
        order_doc = {
            "id": oid, "tenant_id": tenant_id,
            "midtrans_order_id": order_id_mt,
            "customer_phone": phone, "customer_name": sess.get("customer_name", ""),
            "address": txt, "items": cart,
            "subtotal": sess.get("subtotal", total),
            "discount": sess.get("subtotal", total) - total,
            "total": total, "status": "awaiting_payment",
            "created_at": now_utc().isoformat(),
        }
        await db.online_orders.insert_one(order_doc)
        # Create QRIS via Midtrans
        try:
            creds = await get_midtrans_creds(tenant_id)
            payload = {"payment_type": "qris",
                       "transaction_details": {"order_id": order_id_mt, "gross_amount": total},
                       "qris": {"acquirer": "gopay"}}
            async with httpx.AsyncClient(timeout=20) as hc:
                r = await hc.post(f"{midtrans_base(creds['mode'])}/v2/charge",
                                  json=payload, auth=(creds['server_key'], ""),
                                  headers={"Accept": "application/json"})
            data = r.json()
            qr_url = None
            for a in data.get("actions", []):
                if a.get("name") in ("generate-qr-code-v2", "generate-qr-code"):
                    qr_url = a["url"]; break
            await db.online_orders.update_one({"id": oid, "tenant_id": tenant_id},
                {"$set": {"qr_url": qr_url, "midtrans_data": data}})
            # Register in qris_payments for webhook tenant lookup
            await db.qris_payments.insert_one({
                "tenant_id": tenant_id, "order_id": order_id_mt,
                "amount": total, "status": "pending",
                "created_at": now_utc().isoformat(),
            })
            await db.wa_sessions.update_one({"phone": phone, "tenant_id": tenant_id},
                {"$set": {"state": "awaiting_payment", "order_id": oid,
                          "midtrans_order_id": order_id_mt}})
            await wa_send_raw(tenant_id, phone,
                f"📱 Silakan bayar via QRIS berikut sebesar *Rp{total:,}*.\n"
                f"Setelah bayar, kami otomatis proses pesanan Anda. Terima kasih!".replace(",", "."),
                image_url=qr_url)
        except Exception as e:
            logging.exception("QRIS create in WA failed")
            await wa_send_raw(tenant_id, phone,
                f"Mohon maaf, pembayaran QRIS belum bisa dibuat: {str(e)}. Silakan coba lagi nanti.")
        return

    if state == "awaiting_payment":
        # Customer says something while waiting
        await wa_send_raw(tenant_id, phone,
                          "Kami masih menunggu pembayaran QRIS Anda. Ketik *BATAL* untuk membatalkan.")
        return

class WAWebhookIn(BaseModel):
    device: Optional[str] = None
    sender: Optional[str] = None
    message: Optional[str] = None

@api.post("/whatsapp/webhook/{tenant_id}")
async def whatsapp_webhook_tenant(tenant_id: str, request: Request):
    """Tenant-scoped Fonnte webhook. Configure this URL in Fonnte dashboard per tenant."""
    try:
        form = await request.form()
        data = dict(form)
    except Exception:
        data = await request.json()
    sender = str(data.get("sender") or data.get("from") or "").strip()
    message = str(data.get("message") or data.get("text") or "").strip()
    if not sender or not message:
        return {"ok": True, "note": "empty"}
    # Validate tenant exists
    exists = await db.users.find_one({"tenant_id": tenant_id, "role": "owner"})
    if not exists:
        return {"ok": False, "note": "unknown tenant"}
    try:
        asyncio.create_task(bot_reply(tenant_id, sender, message))
    except Exception:
        logging.exception("bot_reply schedule failed")
    return {"ok": True}

@api.post("/whatsapp/webhook")
async def whatsapp_webhook(request: Request):
    """Legacy webhook: routes to the FIRST configured tenant only (kept for backward compat).
    New tenants should use /api/whatsapp/webhook/{tenant_id}.
    """
    try:
        form = await request.form()
        data = dict(form)
    except Exception:
        data = await request.json()
    sender = str(data.get("sender") or data.get("from") or "").strip()
    message = str(data.get("message") or data.get("text") or "").strip()
    if not sender or not message:
        return {"ok": True, "note": "empty"}
    # Route to admin tenant as legacy fallback
    admin = await db.users.find_one({"email": ADMIN_EMAIL.lower()})
    if not admin:
        return {"ok": False, "note": "no admin"}
    tenant_id = admin.get("tenant_id") or admin["id"]
    try:
        asyncio.create_task(bot_reply(tenant_id, sender, message))
    except Exception:
        logging.exception("bot_reply schedule failed")
    return {"ok": True}

# Enhance Midtrans webhook to also handle WA-initiated online orders
@api.post("/whatsapp/simulate-payment/{oid}")
async def wa_simulate_payment(oid: str, user: dict = Depends(get_current_user)):
    """Owner-triggered simulation for demo: mark WA order paid & notify customer."""
    o = await db.online_orders.find_one({"id": oid, "tenant_id": tenant_of(user)})
    if not o:
        raise HTTPException(404, "Order tidak ditemukan")
    await db.online_orders.update_one({"id": oid, "tenant_id": tenant_of(user)},
        {"$set": {"status": "paid", "paid_at": now_utc().isoformat()}})
    await db.wa_sessions.update_one({"phone": o["customer_phone"],
                                      "tenant_id": tenant_of(user)},
        {"$set": {"state": "done"}})
    try:
        await wa_send_raw(tenant_of(user), o["customer_phone"],
            f"✅ *Pembayaran Berhasil!* Pesanan #{oid[:6].upper()} sedang kami proses. Kami akan info saat dikirim. Terima kasih 🙏")
    except Exception:
        pass
    return {"ok": True}

# --- Excel Exports ---
def _xlsx_response(rows: List[dict], sheet_name: str, filename: str, headers: List[str]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    header_fill = PatternFill("solid", fgColor="E85D04")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(h, ""))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'AA'].width = 20
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@api.get("/exports/transactions.xlsx")
async def export_transactions(days: int = 30, user: dict = Depends(get_current_user)):
    cutoff = (now_utc() - timedelta(days=days)).isoformat()
    txs = await db.transactions.find({"tenant_id": tenant_of(user),
                                      "created_at": {"$gte": cutoff}},
                                     {"_id": 0}).sort("created_at", -1).to_list(5000)
    rows = []
    for t in txs:
        items_str = "; ".join([f"{i['qty']}x {i['name']}" for i in t.get("items", [])])
        rows.append({
            "Order ID": t.get("order_id", ""),
            "Tanggal": t.get("created_at", "")[:19].replace("T", " "),
            "Kasir": t.get("cashier", ""),
            "Metode": t.get("payment_method", "").upper(),
            "Status": t.get("status", ""),
            "Subtotal": t.get("subtotal", 0),
            "Diskon": t.get("discount", 0),
            "Total": t.get("total", 0),
            "Laba": t.get("net_profit", 0),
            "Items": items_str,
        })
    return _xlsx_response(rows, "Transaksi", f"transaksi_{days}h.xlsx",
                          ["Order ID", "Tanggal", "Kasir", "Metode", "Status",
                           "Subtotal", "Diskon", "Total", "Laba", "Items"])

@api.get("/exports/inventory.xlsx")
async def export_inventory(user: dict = Depends(get_current_user)):
    prods = await db.products.find({"tenant_id": tenant_of(user)},
                                   {"_id": 0}).sort("name", 1).to_list(2000)
    rows = [{
        "Nama": p["name"], "Kategori": p["category"], "SKU": p.get("sku", ""),
        "Stok": p.get("stock", 0), "Harga Modal": p.get("buy_price", 0),
        "Harga Jual": p.get("sell_price", 0),
        "Nilai Stok (Modal)": p.get("stock", 0) * p.get("buy_price", 0),
    } for p in prods]
    return _xlsx_response(rows, "Inventaris", "inventaris.xlsx",
                          ["Nama", "Kategori", "SKU", "Stok",
                           "Harga Modal", "Harga Jual", "Nilai Stok (Modal)"])

@api.get("/exports/shifts.xlsx")
async def export_shifts(user: dict = Depends(get_current_user)):
    shifts = await db.shifts.find({"tenant_id": tenant_of(user), "status": "closed"},
                                  {"_id": 0}).sort("closed_at", -1).to_list(500)
    rows = []
    for s in shifts:
        t = s.get("totals", {})
        rows.append({
            "Kasir": s.get("user_email", ""),
            "Buka": (s.get("opened_at") or "")[:19].replace("T", " "),
            "Tutup": (s.get("closed_at") or "")[:19].replace("T", " "),
            "Modal Awal": s.get("opening_cash", 0),
            "Kas Tunai": t.get("cash_total", 0),
            "QRIS": t.get("qris_total", 0),
            "Total": t.get("grand_total", 0),
            "Kas Aktual": s.get("closing_cash_actual", 0),
            "Selisih": s.get("discrepancy", 0),
        })
    return _xlsx_response(rows, "Shift", "shift_report.xlsx",
                          ["Kasir", "Buka", "Tutup", "Modal Awal", "Kas Tunai",
                           "QRIS", "Total", "Kas Aktual", "Selisih"])


# --- Seed ---
SEED_PRODUCTS = [
    {"name": "Indomie Goreng", "category": "Mie Instan", "stock": 48, "buy_price": 2800, "sell_price": 3500, "sku": "IDM-GRG-001",
     "image_url": "https://images.unsplash.com/photo-1612927601601-6638404737ce?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Minyak Goreng Sania 1L", "category": "Sembako", "stock": 3, "buy_price": 15500, "sell_price": 18000, "sku": "SNA-MG-1L",
     "image_url": "https://images.unsplash.com/photo-1666694890460-37ec16b0df47?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Kopi Kapal Api Sachet", "category": "Minuman", "stock": 82, "buy_price": 1200, "sell_price": 1500, "sku": "KAP-SCH",
     "image_url": "https://images.unsplash.com/photo-1511920170033-f8396924c348?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Teh Botol Sosro 350ml", "category": "Minuman", "stock": 24, "buy_price": 3500, "sell_price": 4500, "sku": "TBS-350",
     "image_url": "https://images.unsplash.com/photo-1556679343-c7306c1976bc?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Beras Ramos 5kg", "category": "Sembako", "stock": 12, "buy_price": 62000, "sell_price": 72000, "sku": "BRS-5KG",
     "image_url": "https://images.unsplash.com/photo-1586201375761-83865001e31c?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Gula Pasir Gulaku 1kg", "category": "Sembako", "stock": 18, "buy_price": 14000, "sell_price": 16500, "sku": "GLK-1KG",
     "image_url": "https://images.unsplash.com/photo-1519414442781-fbd745c5b497?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Aqua 600ml", "category": "Minuman", "stock": 60, "buy_price": 2500, "sell_price": 3500, "sku": "AQU-600",
     "image_url": "https://images.unsplash.com/photo-1550505095-81378a674395?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Silverqueen Chunky 65g", "category": "Snack", "stock": 15, "buy_price": 13000, "sell_price": 16000, "sku": "SLQ-65",
     "image_url": "https://images.unsplash.com/photo-1548907040-4baa42d10919?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Chitato Sapi Panggang", "category": "Snack", "stock": 22, "buy_price": 7500, "sell_price": 9500, "sku": "CHT-SAPI",
     "image_url": "https://images.unsplash.com/photo-1621939514649-280e2ee25f60?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Roti Tawar Sari Roti", "category": "Roti", "stock": 5, "buy_price": 15000, "sell_price": 18500, "sku": "SR-TAWAR",
     "image_url": "https://images.unsplash.com/photo-1509440159596-0249088772ff?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Susu Ultra Coklat 250ml", "category": "Minuman", "stock": 30, "buy_price": 5500, "sell_price": 7000, "sku": "ULT-CKL",
     "image_url": "https://images.unsplash.com/photo-1550583724-b2692b85b150?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
    {"name": "Sabun Lifebuoy 85g", "category": "Rumah Tangga", "stock": 40, "buy_price": 3500, "sell_price": 4500, "sku": "LB-85",
     "image_url": "https://images.unsplash.com/photo-1585232004423-244e0e6904e3?crop=entropy&cs=srgb&fm=jpg&w=400&q=85"},
]

async def seed_data():
    # indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("tenant_id")
    await db.products.create_index([("tenant_id", 1), ("name", 1)])
    await db.products.create_index([("tenant_id", 1), ("sku", 1)])
    await db.transactions.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.transactions.create_index([("tenant_id", 1), ("shift_id", 1)])
    await db.shifts.create_index([("tenant_id", 1), ("user_id", 1)])
    await db.online_orders.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.online_orders.create_index("midtrans_order_id")
    # Drop legacy unique index on wa_sessions.phone (was global) if present
    try:
        idx = await db.wa_sessions.index_information()
        for name, spec in idx.items():
            keys = spec.get("key", [])
            if [k[0] for k in keys] == ["phone"] and spec.get("unique"):
                await db.wa_sessions.drop_index(name)
    except Exception:
        pass
    await db.wa_sessions.create_index([("tenant_id", 1), ("phone", 1)], unique=True)
    await db.customers.create_index([("tenant_id", 1), ("phone", 1)])
    await db.debt_ledger.create_index([("tenant_id", 1), ("customer_id", 1)])
    await db.expenses.create_index([("tenant_id", 1), ("date", -1)])
    await db.settings.create_index("tenant_id", unique=True, sparse=True)
    await db.qris_payments.create_index("order_id")
    await db.promotions.create_index("tenant_id")

    # admin
    existing = await db.users.find_one({"email": ADMIN_EMAIL.lower()})
    if not existing:
        admin_uid = str(uuid.uuid4())
        await db.users.insert_one({
            "id": admin_uid, "email": ADMIN_EMAIL.lower(),
            "name": "Owner", "role": "owner",
            "tenant_id": admin_uid,
            "password_hash": hash_password(ADMIN_PASSWORD),
            "created_at": now_utc().isoformat(),
        })
        logging.info(f"Seeded admin {ADMIN_EMAIL}")
        admin_tenant = admin_uid
    else:
        # Ensure admin has tenant_id (backfill)
        admin_tenant = existing.get("tenant_id") or existing["id"]
        if not existing.get("tenant_id"):
            await db.users.update_one({"id": existing["id"]},
                                      {"$set": {"tenant_id": admin_tenant}})
        # Update password to match .env
        if not verify_password(ADMIN_PASSWORD, existing["password_hash"]):
            await db.users.update_one({"email": ADMIN_EMAIL.lower()},
                                      {"$set": {"password_hash": hash_password(ADMIN_PASSWORD)}})

    # Migration: backfill tenant_id on legacy docs (from single-tenant era) → assign to admin tenant
    legacy_filter = {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]}
    migrate_collections = ["products", "transactions", "shifts", "customers",
                           "debt_ledger", "expenses", "promotions",
                           "online_orders", "wa_sessions", "qris_payments"]
    for coll in migrate_collections:
        try:
            r = await db[coll].update_many(legacy_filter, {"$set": {"tenant_id": admin_tenant}})
            if r.modified_count:
                logging.info(f"Migrated {r.modified_count} docs in {coll} → tenant {admin_tenant[:8]}")
        except Exception:
            logging.exception(f"Migration failed for {coll}")

    # Migrate other cashier users without tenant_id → admin tenant
    await db.users.update_many({"tenant_id": {"$exists": False}, "role": "cashier"},
                               {"$set": {"tenant_id": admin_tenant}})

    # Migrate legacy settings ({id:"default"}) → tenant-scoped for admin
    legacy_settings = await db.settings.find_one({"id": "default"})
    if legacy_settings and not legacy_settings.get("tenant_id"):
        await db.settings.update_one({"id": "default"},
                                     {"$set": {"tenant_id": admin_tenant},
                                      "$unset": {"id": ""}})

    # products seed (only for admin tenant, and only if that tenant has none)
    if await db.products.count_documents({"tenant_id": admin_tenant}) == 0:
        for p in SEED_PRODUCTS:
            await db.products.insert_one({"id": str(uuid.uuid4()),
                                          "tenant_id": admin_tenant, **p,
                                          "created_at": now_utc().isoformat()})
        logging.info("Seeded sample products for admin tenant")

    # settings default (admin tenant)
    if not await db.settings.find_one({"tenant_id": admin_tenant}):
        await db.settings.insert_one({"tenant_id": admin_tenant,
                                      "store": StoreProfile(name="KasirPintar Demo",
                                                            address="Jl. Sudirman No. 1, Jakarta",
                                                            phone="0812-3456-7890").model_dump(),
                                      "midtrans": {}})

    # dummy transactions for admin if empty
    if await db.transactions.count_documents({"tenant_id": admin_tenant}) == 0:
        products = await db.products.find({"tenant_id": admin_tenant}, {"_id": 0}).to_list(20)
        import random
        for day_offset in range(28, -1, -1):
            date = now_utc() - timedelta(days=day_offset)
            num_tx = random.randint(2, 6)
            for _ in range(num_tx):
                num_items = random.randint(1, 3)
                items = []
                subtotal = 0
                profit = 0
                for _ in range(num_items):
                    pr = random.choice(products)
                    qty = random.randint(1, 3)
                    items.append({"product_id": pr["id"], "name": pr["name"],
                                  "qty": qty, "price": pr["sell_price"],
                                  "buy_price": pr["buy_price"], "discount": 0})
                    subtotal += pr["sell_price"] * qty
                    profit += (pr["sell_price"] - pr["buy_price"]) * qty
                tid = str(uuid.uuid4())
                await db.transactions.insert_one({
                    "id": tid, "tenant_id": admin_tenant,
                    "order_id": f"TRX-{date.strftime('%y%m%d%H%M%S')}-{tid[:4].upper()}",
                    "items": items, "subtotal": subtotal, "discount": 0, "tax": 0,
                    "total": subtotal, "payment_method": random.choice(["cash", "qris"]),
                    "cash_tendered": subtotal, "change": 0,
                    "status": "completed", "net_profit": profit,
                    "cashier": ADMIN_EMAIL,
                    "created_at": date.isoformat(),
                })
        logging.info("Seeded dummy transactions for admin tenant")

@app.on_event("startup")
async def startup():
    await seed_data()

app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origin_regex=".*",
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown():
    client.close()
